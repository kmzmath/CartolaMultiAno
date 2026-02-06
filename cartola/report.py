from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional, Any, TYPE_CHECKING

import numpy as np
import pandas as pd

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from .io import PoissonModel, MatchOdds, safe_int

if TYPE_CHECKING:
    from .models import PositionModels

from .config import STATUS_ORD, SHEET_BY_POS, POSICOES

from pathlib import Path

@dataclass
class PlayerPrediction:
    atleta_id: int
    apelido: str
    posicao_id: int
    posicao: str
    clube_id: int
    clube: str
    opponent_id: int
    opponent: str
    is_home: bool
    preco: float
    rank_score: float
    predicted_score: float
    pred_p10: float
    pred_p90: float
    interval_width: float
    status: str
    recommendation_score: float
    features: Dict[str, float] = field(default_factory=dict)


def print_backtest(results: Dict[str, Any], k_pos: int = 5):
    print("\n" + "=" * 100)
    print(f" 📊 BACKTEST (POSICIONAL) - TOP {k_pos} POR POSIÇÃO (com lift/percentil)")
    print("=" * 100)

    if not results:
        print("Sem resultados (dados insuficientes).")
        return

    print(f"\n📈 {results.get('n_rodadas', 0)} rodadas | {results.get('total', 0)} previsões")

    overall = results.get("overall_weighted", {})
    by_pos = results.get("by_pos", {})

    header = f"{'Métrica':<30} {'Média':>12} {'Std':>10} {'Min':>10} {'Max':>10}"

    main_keys = [
        f"ndcg@{k_pos}",
        f"hit_rate@{k_pos}",
        f"top{k_pos}_overlap",
        f"top{k_pos}_pct",
        f"lift@{k_pos}",
        f"regret@{k_pos}",
        f"pick_percentile_mean@{k_pos}",
        f"pick_percentile_median@{k_pos}",
        f"pick_percentile_min@{k_pos}",
        f"pct_picks_above_mean@{k_pos}",
        "mae",
        "rmse",
    ]

    print("\n--- Geral (ponderado por posição) ---")
    print(header)
    print("-" * len(header))
    for k in main_keys:
        if k in overall:
            m = overall[k]
            print(f"{k:<30} {m['mean']:>12.3f} {m['std']:>10.3f} {m['min']:>10.3f} {m['max']:>10.3f}")

    if by_pos:
        print("\n--- Por posição ---")
        for pos_id in sorted(by_pos.keys()):
            pname = POSICOES.get(pos_id, str(pos_id))
            mp = by_pos[pos_id]
            print(f"\n[{pname} | id={pos_id}]")
            print(header)
            print("-" * len(header))
            for k in main_keys:
                if k in mp:
                    m = mp[k]
                    print(f"{k:<30} {m['mean']:>12.3f} {m['std']:>10.3f} {m['min']:>10.3f} {m['max']:>10.3f}")

def _ensemble_importance(ens, top_n: int = 40, kind: str = "ranker"):
    """
    kind:
      - "ranker" -> importance do LGBMRanker (rank_score)
      - "score"  -> importance do LGBMRegressor (previsao_pontos)
    Retorna lista [(feature, importance_normalizada), ...] com média no ensemble.
    """
    if not ens:
        return []

    # Define colunas "base" para alinhar (evita bagunça se alguma seed variar)
    base_cols = getattr(ens[0], "feature_columns", None) or []
    if not base_cols:
        return []

    acc = np.zeros(len(base_cols), dtype=float)
    used = 0

    for m in ens:
        cols = getattr(m, "feature_columns", None) or []
        model_obj = getattr(m, "ranker", None) if kind == "ranker" else getattr(m, "regressor", None)
        if model_obj is None or not cols:
            continue

        imp = getattr(model_obj, "feature_importances_", None)
        if imp is None:
            continue

        imp = np.asarray(imp, dtype=float)

        # Alinha importance nas colunas base (caso a ordem/colunas difiram)
        if cols != base_cols:
            idx = {f: i for i, f in enumerate(cols)}
            aligned = np.zeros(len(base_cols), dtype=float)
            for i, f in enumerate(base_cols):
                j = idx.get(f)
                if j is not None and j < len(imp):
                    aligned[i] = imp[j]
            imp = aligned

        s = float(imp.sum())
        if s <= 0:
            continue

        acc += imp / s
        used += 1

    if used == 0:
        return []

    acc /= used
    items = list(zip(base_cols, acc))
    items.sort(key=lambda x: x[1], reverse=True)
    return items[:top_n]


def importance_to_df(models: "PositionModels", top_n: int = 40, kind: str = "ranker", posicao_id: Optional[int] = None) -> pd.DataFrame:
    """
    posicao_id=None -> GLOBAL (ensemble global)
    posicao_id=int  -> importance do ensemble daquela posição (cai no global se não existir)
    """
    if posicao_id is None:
        ens = getattr(models, "global_models", None) or []
    else:
        ens = getattr(models, "models", {}).get(posicao_id) or []
        if not ens:
            ens = getattr(models, "global_models", None) or []

    imp = _ensemble_importance(ens, top_n=top_n, kind=kind)
    return pd.DataFrame([{"feature": f, "importance": v} for f, v in imp])


def print_importance(models: "PositionModels", top_n: int = 25):
    def _print_block(title: str, df: pd.DataFrame):
        print("\n" + "=" * 100)
        print(title)
        print("=" * 100)

        if df is None or df.empty:
            print("Sem importance disponível.")
            return

        print(f"\n{'Feature':<40} {'Importância':>15}")
        print("-" * 55)
        for _, row in df.iterrows():
            feat = str(row["feature"])
            v = float(row["importance"])
            bar = "█" * int(v * 120)
            print(f"{feat[:39]:<40} {v:>12.4f}  {bar}")

    # GLOBAL
    _print_block(" 🎯 FEATURE IMPORTANCE - GLOBAL (RANKER / rank_score)", importance_to_df(models, top_n=top_n, kind="ranker", posicao_id=None))
    _print_block(" 🎯 FEATURE IMPORTANCE - GLOBAL (REGRESSOR / previsao_pontos)", importance_to_df(models, top_n=top_n, kind="score", posicao_id=None))

    # POR POSIÇÃO
    for pos_id in sorted(POSICOES.keys()):
        pname = POSICOES.get(pos_id, str(pos_id))
        df_r = importance_to_df(models, top_n=top_n, kind="ranker", posicao_id=pos_id)
        df_s = importance_to_df(models, top_n=top_n, kind="score", posicao_id=pos_id)

        _print_block(f" 🎯 FEATURE IMPORTANCE - {pname} (RANKER / rank_score)", df_r)
        _print_block(f" 🎯 FEATURE IMPORTANCE - {pname} (REGRESSOR / previsao_pontos)", df_s)


def print_matches(odds_by_clube: Dict[int, MatchOdds]):
    print("\n" + "=" * 100)
    print(" 📊 PARTIDAS (odds → 1X2 → Poisson xG)")
    print("=" * 100)
    shown = set()
    for m in odds_by_clube.values():
        key = (m.home_id, m.away_id)
        if key in shown:
            continue
        shown.add(key)
        pcs_home = PoissonModel.poisson_prob(m.lambda_away, 0) * 100
        pcs_away = PoissonModel.poisson_prob(m.lambda_home, 0) * 100
        print(f"\n{m.home_team} vs {m.away_team}")
        print(f"   Prob: {m.p_home * 100:.1f}% / {m.p_draw * 100:.1f}% / {m.p_away * 100:.1f}%")
        print(
            f"   xG: {m.lambda_home:.2f} - {m.lambda_away:.2f} | "
            f"P(CS): {m.home_team}={pcs_home:.0f}% / {m.away_team}={pcs_away:.0f}%"
        )


def _autosize_and_format_sheet(ws, df: pd.DataFrame):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), *(len(str(x)) for x in df[col].astype(str).values[:200]))
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, max_len + 2), 45)

    fmt_2 = "0.00"
    fmt_1 = "0.0"
    fmt_pct = "0.0%"

    colmap = {c: i+1 for i, c in enumerate(df.columns)}
    def set_fmt(colname, fmt):
        if colname not in colmap:
            return
        j = colmap[colname]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=j).number_format = fmt

    for c in ["previsao_pontos", "p10", "p90", "intervalo", "rank_score",
              "team_xG", "opp_xG", "player_mean", "player_last5_mean", "team_avg", "pos_vs_opp_mean"]:
        set_fmt(c, fmt_2)

    for c in ["preco"]:
        set_fmt(c, fmt_1)

    for c in ["p_clean_sheet", "p_team_scores_2plus"]:
        set_fmt(c, fmt_pct)

def safe_int(v, default=0):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return int(v)
    except (ValueError, TypeError):
        return default

def preds_to_df(preds: List[PlayerPrediction]) -> pd.DataFrame:
    rows = []
    for p in preds:
        rows.append({
            "posicao_id": p.posicao_id,
            "posicao": p.posicao,
            "status": p.status,
            "atleta_id": p.atleta_id,
            "apelido": p.apelido,
            "clube": p.clube,
            "adversario": p.opponent,
            "mandante": p.is_home,
            "previsao_pontos": p.predicted_score,
            "p10": p.pred_p10,
            "p90": p.pred_p90,
            "intervalo": p.interval_width,
            "rank_score": p.rank_score,
            "preco": p.preco,
            "team_xG": p.features.get("team_xG", 0.0),
            "opp_xG": p.features.get("opp_xG", 0.0),
            "p_clean_sheet": p.features.get("p_clean_sheet", 0.0),
            "p_team_scores_2plus": p.features.get("p_team_scores_2plus", 0.0),
            "tendencia": p.features.get("player_trend", 0.0),
            "player_mean": p.features.get("player_mean", 0.0),
            "player_last5_mean": p.features.get("player_last5_mean", 0.0),
            "team_avg": p.features.get("team_avg", 0.0),
            "pos_vs_opp_mean": p.features.get("pos_vs_opp_mean", 0.0),
            "jogos": safe_int(p.features.get("player_games"), 0),
        })

    df = pd.DataFrame(rows)

    df["status_ord"] = df["status"].map(lambda s: STATUS_ORD.get(s, 9)).astype(int)
    df = df.sort_values(
        ["posicao_id", "status_ord", "previsao_pontos"],
        ascending=[True, True, False]
    ).drop(columns=["status_ord"])

    col_order = [
        "posicao_id","posicao","status","apelido","clube","adversario","mandante",
        "previsao_pontos","p10","p90","intervalo","rank_score","preco",
        "team_xG","opp_xG","p_clean_sheet","p_team_scores_2plus","tendencia",
        "player_mean","player_last5_mean","team_avg","pos_vs_opp_mean","jogos","atleta_id"
    ]
    col_order = [c for c in col_order if c in df.columns]
    df = df[col_order]

    return df


def matches_to_df(odds_by_clube: Dict[int, MatchOdds]) -> pd.DataFrame:
    rows = []
    seen = set()
    for m in odds_by_clube.values():
        key = (m.home_id, m.away_id)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "casa": m.home_team,
            "fora": m.away_team,
            "odd_casa": m.odd_home,
            "odd_empate": m.odd_draw,
            "odd_fora": m.odd_away,
            "p_casa": m.p_home,
            "p_empate": m.p_draw,
            "p_fora": m.p_away,
            "xG_casa": m.lambda_home,
            "xG_fora": m.lambda_away,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["casa", "fora"])
    return df


def save_excel(
    preds: List[PlayerPrediction],
    odds_by_clube: Dict[int, MatchOdds],
    models: "PositionModels",
    path: str,
    top_n_resumo: int = 25,
):
    # garante diretório
    out_path = Path(path)
    if out_path.parent and str(out_path.parent) not in ("", "."):
        out_path.parent.mkdir(parents=True, exist_ok=True)

    df_all = preds_to_df(preds)

    # =========================
    # Resumo (top geral + top por posição)
    # =========================
    resumo_parts: List[pd.DataFrame] = []

    if not df_all.empty:
        top_score = (
            df_all.sort_values(["previsao_pontos"], ascending=False)
            .head(top_n_resumo)
            .copy()
        )
        top_score.insert(0, "lista", f"TOP {top_n_resumo} - Pontos")
        resumo_parts.append(top_score)

        for pos_id, pos_name in POSICOES.items():
            d = df_all[df_all["posicao_id"] == pos_id].copy()
            if d.empty:
                continue
            t = d.sort_values(["previsao_pontos"], ascending=False).head(top_n_resumo).copy()
            t.insert(0, "lista", f"{pos_name} - TOP {top_n_resumo} (Pontos)")
            resumo_parts.append(t)

    df_resumo = pd.concat(resumo_parts, ignore_index=True) if resumo_parts else pd.DataFrame()

    # =========================
    # Partidas
    # =========================
    df_matches = matches_to_df(odds_by_clube)

    # =========================
    # Importance (global + por posição)
    # =========================
    df_imp_rank_global = importance_to_df(models, top_n=40, kind="ranker", posicao_id=None)
    df_imp_score_global = importance_to_df(models, top_n=40, kind="score", posicao_id=None)

    sheet_by_pos = SHEET_BY_POS

    # =========================
    # Escrita do Excel
    # =========================
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        if not df_resumo.empty:
            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

        # Base
        df_all.to_excel(writer, sheet_name="Jogadores", index=False)

        # Abas por posição (jogadores)
        for pos_id, sheet_name in sheet_by_pos.items():
            d = df_all[df_all["posicao_id"] == pos_id].copy()
            if not d.empty:
                df_sheet = d
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

        # Partidas
        if not df_matches.empty:
            df_matches.to_excel(writer, sheet_name="Partidas", index=False)

        # Importance (global)
        if df_imp_rank_global is not None and not df_imp_rank_global.empty:
            df_imp_rank_global.to_excel(writer, sheet_name="Importance_Ranker", index=False)

        if df_imp_score_global is not None and not df_imp_score_global.empty:
            df_imp_score_global.to_excel(writer, sheet_name="Importance_Score", index=False)

        # Importance (por posição)
        for pos_id, sheet_name in sheet_by_pos.items():
            df_r = importance_to_df(models, top_n=40, kind="ranker", posicao_id=pos_id)
            df_s = importance_to_df(models, top_n=40, kind="score", posicao_id=pos_id)

            if df_r is not None and not df_r.empty:
                name_r = f"ImpR_{sheet_name}"
                df_r.to_excel(writer, sheet_name=name_r[:31], index=False)  # Excel <= 31 chars

            if df_s is not None and not df_s.empty:
                name_s = f"ImpS_{sheet_name}"
                df_s.to_excel(writer, sheet_name=name_s[:31], index=False)

    # =========================
    # Autoformatação
    # =========================
    from openpyxl import load_workbook

    wb = load_workbook(str(out_path))
    for sh in wb.sheetnames:
        ws = wb[sh]
        values = list(ws.values)
        if not values or len(values) < 2:
            continue
        cols = list(values[0])
        data = values[1:]
        df_tmp = pd.DataFrame(data, columns=cols)
        _autosize_and_format_sheet(ws, df_tmp)

    wb.save(str(out_path))